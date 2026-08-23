"""Import the 2025 listed-company regulation workbook into static site data."""

from __future__ import annotations

import argparse
import json
from datetime import datetime
from pathlib import Path

import openpyxl

VERIFIED_AT = "2026-08-23"
VERIFIED_LINKS = {
    "以上市公司质量为导向的保荐机构执业质量评价实施办法（试行）（2025年修订）": {
        "officialUrl": "https://www.sse.com.cn/lawandrules/sselawsrules2025/stocks/review/firstepisode/c/c_20250514_10778839.shtml",
        "verifiedAt": VERIFIED_AT,
    },
    "北京证券交易所股票上市规则": {
        "officialUrl": "https://www.bse.cn/cxjg_list/200025638.html",
        "verifiedAt": VERIFIED_AT,
    },
    "北京证券交易所上市公司持续监管指引第13号——股份变动管理": {
        "officialUrl": "https://www.bse.cn/cxjg_list/200025609.html",
        "verifiedAt": VERIFIED_AT,
    },
}


def clean(value):
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, str):
        return " ".join(value.split())
    return value


def import_workbook(source: Path):
    workbook = openpyxl.load_workbook(source, read_only=True, data_only=True)
    records = []
    counts = {}

    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        category = "未分类"
        sheet_records = 0

        for row in worksheet.iter_rows(values_only=True):
            number, published_at, document_number, name = (list(row) + [None] * 4)[:4]
            number = clean(number)
            published_at = clean(published_at)
            document_number = clean(document_number)
            name = clean(name)

            if isinstance(number, str) and number and not published_at and not document_number and not name:
                category = number
                continue
            if not isinstance(number, int) or not name:
                continue

            link = VERIFIED_LINKS.get(name)
            records.append(
                {
                    "id": f"{sheet_name}-{number}-{sheet_records + 1}",
                    "serialNumber": number,
                    "publishedAt": published_at,
                    "documentNumber": document_number,
                    "name": name,
                    "authority": sheet_name,
                    "category": category,
                    "officialUrl": link["officialUrl"] if link else None,
                    "linkLabel": "官方原文" if link else "待补充链接",
                    "linkStatus": "verified" if link else "pending",
                    "verifiedAt": link["verifiedAt"] if link else None,
                }
            )
            sheet_records += 1

        counts[sheet_name] = sheet_records

    records.sort(key=lambda item: (item["publishedAt"] or "0000-00-00", item["authority"], item["serialNumber"]), reverse=True)
    verified_count = sum(1 for item in records if item["linkStatus"] == "verified")
    years = sorted({item["publishedAt"][:4] for item in records if item["publishedAt"]}, reverse=True)
    return {
        "version": "2025.1",
        "source": "2025年上市公司法规汇编.xlsx",
        "importedAt": VERIFIED_AT,
        "recordCount": len(records),
        "verifiedLinkCount": verified_count,
        "pendingLinkCount": len(records) - verified_count,
        "years": years,
        "authorityCounts": counts,
        "records": records,
    }


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("source", type=Path)
    parser.add_argument("output", type=Path)
    args = parser.parse_args()
    payload = import_workbook(args.source)
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps({key: payload[key] for key in ("recordCount", "verifiedLinkCount", "pendingLinkCount", "authorityCounts")}, ensure_ascii=False))


if __name__ == "__main__":
    main()
